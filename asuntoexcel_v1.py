from openpyxl import Workbook
from openpyxl.styles import Font

# Luo uusi Excel-työkirja
wb = Workbook()
ws = wb.active
ws.title = "Asumislaskuri"

# Syöttöarvot
inputs = {
    "Asunnon hinta (€)": 480000,
    "Asunnon koko (m2)": 60,
    "Nykyinen varallisuus (€)": 150000,
    "Korko (%)": 3.5,
    "Yhtiövastike (€/m2)": 5,
    "Vuokra alussa (€/kk)": 1600,
    "Vuokran nousu (%/v)": 2,
    "Asunnon arvonnousu (%/v)": 1,
    "Putkiremontti (€/m2, 20v kohdalla)": 1000,
    "Sijoituksen tuotto (%/v)": 4,
    "Muut kulut/vuosi (€)": 200
}

ws["A1"] = "Syöttöarvot"
ws["A1"].font = Font(bold=True)
row = 2
input_refs = {}
for key, value in inputs.items():
    ws[f"A{row}"] = key
    cell = f"B{row}"
    ws[cell] = value
    input_refs[key] = cell
    row += 1

# Tulostaulukko
ws["A15"] = "Tulokset"
ws["A15"].font = Font(bold=True)
headers = ["Tieto", "5 vuotta", "10 vuotta", "20 vuotta"]
for col_num, header in enumerate(headers, 1):
    ws.cell(row=16, column=col_num, value=header).font = Font(bold=True)

# Määritetään laskettavat rivit
labels = [
    "Omistusasujan varallisuus (€)",
    "Vuokralaisen varallisuus (€)",
    "Erotus (€)"
]

durations = [5, 10, 20]
start_row = 17

# Soluviitteet
price = input_refs["Asunnon hinta (€)"]
size = input_refs["Asunnon koko (m2)"]
capital = input_refs["Nykyinen varallisuus (€)"]
interest = input_refs["Korko (%)"]
maintenance = input_refs["Yhtiövastike (€/m2)"]
rent_start = input_refs["Vuokra alussa (€/kk)"]
rent_increase = input_refs["Vuokran nousu (%/v)"]
appreciation = input_refs["Asunnon arvonnousu (%/v)"]
renovation = input_refs["Putkiremontti (€/m2, 20v kohdalla)"]
investment = input_refs["Sijoituksen tuotto (%/v)"]
other_costs = input_refs["Muut kulut/vuosi (€)"]

# Kirjoitetaan kaavat laskentataulukkoon
for i, label in enumerate(labels):
    ws[f"A{start_row + i}"] = label
    for j, years in enumerate(durations):
        col = j + 2
        cell = ws.cell(row=start_row + i, column=col)
        if label == "Omistusasujan varallisuus (€)":
            fv = f"{price}*(1+{appreciation}/100)^{years}"
            base_loan = f"({price}-{capital})"
            monthly_interest = f"{interest}/100/12"
            n_months = years * 12
            annuity = f"={base_loan}*{monthly_interest}/(1-(1+{monthly_interest})^-{n_months})"
            remaining_loan = f"({base_loan})*((1+{monthly_interest})^({25*12})-(1+{monthly_interest})^{n_months})/((1+{monthly_interest})^{25*12}-1)"
            net_equity = f"={fv}-({remaining_loan})"
            cell.value = net_equity
        elif label == "Vuokralaisen varallisuus (€)":
            rent_paid = f"=12*{rent_start}*(((1+{rent_increase}/100)^{years}-1)/({rent_increase}/100))"
            invested = f"{capital}*(1+{investment}/100)^{years}"
            net_renter = f"={invested}-({rent_paid})"
            cell.value = net_renter
        elif label == "Erotus (€)":
            owner_cell = ws.cell(row=start_row, column=col).coordinate
            renter_cell = ws.cell(row=start_row+1, column=col).coordinate
            cell.value = f"={owner_cell}-{renter_cell}"

# Tallennetaan tiedosto
excel_path = "/mnt/data/Asumislaskuri_varallisuusvertailu.xlsx"
wb.save(excel_path)

excel_path
