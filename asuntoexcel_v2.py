# Uusitaan tiedoston luonti istunnon nollauksen jälkeen
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Luo uusi työkirja
wb = Workbook()
ws = wb.active
ws.title = "Asumislaskuri"

# Syöttöarvot
inputs = {
    "Asunnon hinta (€)": 480000,
    "Asunnon koko (m2)": 60,
    "Lainan osuus (%)": 85,
    "Korko (%)": 3.5,
    "Yhtiövastike (€/m2)": 5,
    "Vuokra alussa (€/kk)": 1600,
    "Vuokran nousu (%/v)": 2,
    "Asunnon arvonnousu (%/v)": 1,
    "Putkiremontti (€/m2, 20v kohdalla)": 1000,
    "Sijoituksen tuotto (%/v)": 4,
    "Muut kulut/vuosi (€)": 200
}

ws["A1"] = "Syöttöarvot"
ws["A1"].font = Font(bold=True)
row = 2
input_refs = {}
for key, value in inputs.items():
    ws[f"A{row}"] = key
    cell = f"B{row}"
    ws[cell] = value
    input_refs[key] = cell
    row += 1

# Tulostaulukon otsikot
ws["A15"] = "Tulokset"
ws["A15"].font = Font(bold=True)
headers = ["Tieto", "5 vuotta", "10 vuotta", "20 vuotta"]
for col_num, header in enumerate(headers, 1):
    ws.cell(row=16, column=col_num, value=header).font = Font(bold=True)

# Kaavatulokset
price = input_refs["Asunnon hinta (€)"]
size = input_refs["Asunnon koko (m2)"]
ltv = input_refs["Lainan osuus (%)"]
interest = input_refs["Korko (%)"]
maintenance = input_refs["Yhtiövastike (€/m2)"]
rent_start = input_refs["Vuokra alussa (€/kk)"]
rent_increase = input_refs["Vuokran nousu (%/v)"]
appreciation = input_refs["Asunnon arvonnousu (%/v)"]
renovation = input_refs["Putkiremontti (€/m2, 20v kohdalla)"]
investment = input_refs["Sijoituksen tuotto (%/v)"]
other_costs = input_refs["Muut kulut/vuosi (€)"]

labels = [
    "Omistusasuminen: netto (€)",
    "Vuokra-asuminen: netto (€)",
    "Arvioitu asunnon arvo (€)",
    "Maksettu vuokra yhteensä (€)",
    "Sijoituksen arvo (€)"
]

durations = [5, 10, 20]
start_row = 17

for i, label in enumerate(labels):
    ws[f"A{start_row + i}"] = label
    for j, years in enumerate(durations):
        col_letter = get_column_letter(j + 2)
        row = start_row + i

        if label == "Arvioitu asunnon arvo (€)":
            formula = f"={price}*(1+{appreciation}/100)^{years}"
        elif label == "Maksettu vuokra yhteensä (€)":
            formula = f"=12*{rent_start}*(((1+{rent_increase}/100)^{years}-1)/({rent_increase}/100))"
        elif label == "Sijoituksen arvo (€)":
            formula = f"=({price}*(1-{ltv}/100))*(1+{investment}/100)^{years}"
        elif label == "Omistusasuminen: netto (€)":
            base_loan = f"{price}*{ltv}/100"
            monthly_interest = f"{interest}/100/12"
            n_months = years * 12
            annuity = f"={base_loan}*{monthly_interest}/(1-(1+{monthly_interest})^-{n_months})"
            total_loan = f"({annuity})*{n_months}"
            maintenance_costs = f"={maintenance}*{size}*12*{years}"
            yearly_other = f"{other_costs}*{years}"
            renovation_cost = f"=IF({years}=20,{renovation}*{size},0)"
            value_gain = f"={price}*(1+{appreciation}/100)^{years}"
            remaining_equity = f"{value_gain}-({base_loan})"
            formula = f"={total_loan}+{maintenance_costs}+{yearly_other}+{renovation_cost}-{remaining_equity}"
        elif label == "Vuokra-asuminen: netto (€)":
            total_rent = f"=12*{rent_start}*(((1+{rent_increase}/100)^{years}-1)/({rent_increase}/100))"
            invested = f"=({price}*(1-{ltv}/100))*(1+{investment}/100)^{years}"
            formula = f"={total_rent}-({invested}-{price}*(1-{ltv}/100))"
        else:
            formula = ""

        ws[f"{col_letter}{row}"] = formula

# Kaavio
chart = LineChart()
chart.title = "Omistus- vs vuokra-asuminen"
chart.style = 13
chart.y_axis.title = "Kustannus (€)"
chart.x_axis.title = "Ajanjakso (vuotta)"

data = Reference(ws, min_col=2, max_col=4, min_row=17, max_row=18)
categories = Reference(ws, min_col=2, max_col=4, min_row=16)
chart.add_data(data, titles_from_data=False)
chart.set_categories(categories)
chart.series[0].title = SeriesLabel(v="Omistusasuminen")
chart.series[1].title = SeriesLabel(v="Vuokra-asuminen")

ws.add_chart(chart, "F5")

# Tallennetaan uusi tiedosto
excel_path = "/mnt/data/Asumislaskuri_kaaviolla_uudelleen.xlsx"
wb.save(excel_path)

excel_path
